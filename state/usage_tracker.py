from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

import structlog

logger = structlog.get_logger()

USAGE_FILE = Path(__file__).parent / "runs" / "usage_log.xlsx"

# Pricing per token — input, input_cached, output
PRICING = {
    "gpt-4o-mini": {"input": 0.15 / 1_000_000, "input_cached": 0.075 / 1_000_000,  "output": 0.60 / 1_000_000},
    "gpt-4o":      {"input": 2.50 / 1_000_000, "input_cached": 1.25 / 1_000_000,   "output": 10.0 / 1_000_000},
    "gpt-4.1":     {"input": 2.00 / 1_000_000, "input_cached": 0.50 / 1_000_000,   "output": 8.00 / 1_000_000},
    "gpt-5":       {"input": 1.25 / 1_000_000, "input_cached": 0.125 / 1_000_000,  "output": 10.0 / 1_000_000},
    "gpt-5.1":     {"input": 1.25 / 1_000_000, "input_cached": 0.125 / 1_000_000,  "output": 10.0 / 1_000_000},
    "openai/gpt-oss-120b": {"input": 0.039 / 1_000_000, "input_cached": 0.039 / 1_000_000, "output": 0.19 / 1_000_000},
    # Groq (free tier)
    "groq/llama-3.3-70b-versatile":  {"input": 0.59 / 1_000_000, "input_cached": 0.59 / 1_000_000, "output": 0.79 / 1_000_000},
    "groq/llama-3.1-8b-instant":     {"input": 0.05 / 1_000_000, "input_cached": 0.05 / 1_000_000, "output": 0.08 / 1_000_000},
    # OpenRouter (free models — $0)
    "openrouter/google/gemini-2.5-flash":              {"input": 0, "input_cached": 0, "output": 0},
    "openrouter/deepseek/deepseek-v3.2-20251201":      {"input": 0, "input_cached": 0, "output": 0},
}

USD_TO_INR = 91.04  # As of 27 Feb 2026 — update as needed

HEADERS = [
    "Timestamp",
    "Run Mode",
    "Model",
    "Turns Used",
    "Input Tokens",
    "Cached Tokens",
    "Uncached Tokens",
    "Cache Hit %",
    "Output Tokens",
    "Total Tokens",
    "Uncached Cost ($)",
    "Cached Cost ($)",
    "Output Cost ($)",
    "Real Cost ($)",
    "Real Cost (₹)",
    "No-Cache Cost ($)",
    "Savings ($)",
    "Savings %",
    "Cumulative Real ($)",
    "Cumulative Real (₹)",
    "Duration (sec)",
    "Status",
    "Notes",
]


def _get_or_create_workbook() -> tuple[Workbook, Path]:
    USAGE_FILE.parent.mkdir(parents=True, exist_ok=True)
    if USAGE_FILE.exists():
        wb = load_workbook(USAGE_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Usage Log"
        ws.append(HEADERS)
        for col_num, header in enumerate(HEADERS, 1):
            col_letter = chr(64 + col_num) if col_num <= 26 else "A"
            ws.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        wb.save(USAGE_FILE)
        logger.info("usage_log_created", path=str(USAGE_FILE))
    return wb, USAGE_FILE


def log_usage(
    run_mode: str,
    model: str,
    turns_used: int,
    input_tokens: int,
    output_tokens: int,
    cached_tokens: int = 0,
    duration_sec: float = 0.0,
    status: str = "completed",
    notes: str = "",
) -> dict:
    prices = PRICING.get(model, PRICING["gpt-4o-mini"])

    uncached_tokens = input_tokens - cached_tokens
    total_tokens = input_tokens + output_tokens
    cache_hit_pct = (cached_tokens / input_tokens * 100) if input_tokens > 0 else 0

    # Real cost (with caching discount)
    uncached_cost = uncached_tokens * prices["input"]
    cached_cost = cached_tokens * prices["input_cached"]
    output_cost = output_tokens * prices["output"]
    real_cost = uncached_cost + cached_cost + output_cost

    # What it would have cost without caching
    nocache_cost = input_tokens * prices["input"] + output_cost
    savings = nocache_cost - real_cost
    savings_pct = (savings / nocache_cost * 100) if nocache_cost > 0 else 0

    # INR conversion — per this run only
    real_cost_inr = real_cost * USD_TO_INR

    wb, path = _get_or_create_workbook()
    ws = wb.active

    # Calculate cumulative real cost from previous rows
    cumulative = real_cost
    # Find the "Cumulative Real ($)" column index
    cum_col_idx = None
    for idx, cell in enumerate(ws[1], 0):
        if cell.value == "Cumulative Real ($)":
            cum_col_idx = idx
            break

    if cum_col_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[cum_col_idx] is not None:
                cumulative = row[cum_col_idx] + real_cost

    row_data = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        run_mode,
        model,
        turns_used,
        input_tokens,
        cached_tokens,
        uncached_tokens,
        round(cache_hit_pct, 1),
        output_tokens,
        total_tokens,
        round(uncached_cost, 6),
        round(cached_cost, 6),
        round(output_cost, 6),
        round(real_cost, 6),
        round(real_cost_inr, 4),
        round(nocache_cost, 6),
        round(savings, 6),
        round(savings_pct, 1),
        round(cumulative, 6),
        round(cumulative * USD_TO_INR, 4),
        round(duration_sec, 2),
        status,
        notes,
    ]
    ws.append(row_data)
    wb.save(path)

    summary = {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "cached_tokens": cached_tokens,
        "uncached_tokens": uncached_tokens,
        "total_tokens": total_tokens,
        "cache_hit_pct": round(cache_hit_pct, 1),
        "real_cost": round(real_cost, 6),
        "real_cost_inr": round(real_cost_inr, 4),
        "nocache_cost": round(nocache_cost, 6),
        "savings": round(savings, 6),
        "savings_pct": round(savings_pct, 1),
        "cumulative_cost": round(cumulative, 6),
    }

    logger.info("usage_logged", model=model,
                real_cost=f"${real_cost:.6f}",
                nocache=f"${nocache_cost:.6f}",
                savings=f"${savings:.6f} ({savings_pct:.1f}%)",
                cache_hit=f"{cache_hit_pct:.1f}%",
                cost_inr=f"₹{real_cost_inr:.4f}",
                cumulative=f"${cumulative:.6f}")
    return summary
