"""
Sync script — reads existing Excel + JSON + txt files and loads into SQLite.
This is READ-ONLY: it never modifies existing files. Safe to re-run (idempotent).

Usage:
    python state/sync.py              # Sync everything
    python state/sync.py --stats      # Show what's in the DB after sync
"""

import re
import sys
import json
import sqlite3
from pathlib import Path

# openpyxl is already in your requirements (used by usage_tracker.py)
from openpyxl import load_workbook


# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent.parent                      # MCP_agent/
DB_PATH = Path(__file__).parent / "agent_analytics.db"   # state/agent_analytics.db
USAGE_XLSX = Path(__file__).parent / "runs" / "usage_log.xlsx"

# All directories that contain output/turn files
RUN_DIRS = [
    ROOT / "state" / "runs",           # v1 outputs
    ROOT / "version_2" / "runs",       # v2 outputs (root)
    ROOT / "version_2" / "runs" / "pass1",  # v2 Pass 1
    ROOT / "version_2" / "runs" / "pass2",  # v2 Pass 2
]

# StateTracker JSON files
STATE_DIR = ROOT / "state" / "runs"


# ── Database connection ───────────────────────────────────────────────────────

def get_db() -> sqlite3.Connection:
    """Connect to SQLite. Creates tables if DB doesn't exist yet."""
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")    # better concurrent read performance
    db.execute("PRAGMA foreign_keys=ON")
    return db


# ── 1. Sync runs from usage_log.xlsx ─────────────────────────────────────────

def sync_runs_from_excel(db: sqlite3.Connection) -> int:
    """Read usage_log.xlsx and insert each row into the runs table.

    Skips rows that already exist (matched by timestamp + mode + model).
    Returns count of new rows inserted.
    """
    if not USAGE_XLSX.exists():
        print(f"  [skip] {USAGE_XLSX} not found")
        return 0

    wb = load_workbook(USAGE_XLSX, read_only=True)
    ws = wb.active

    # Read header row to get column positions
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(name: str) -> int:
        """Get column index by header name."""
        return headers.index(name)

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # skip empty rows
            continue

        timestamp = str(row[col("Timestamp")])
        mode = row[col("Run Mode")]
        model = row[col("Model")]

        # Build a run_id from timestamp: "2026-03-11 16:45:18" -> "20260311_164518"
        run_id = re.sub(r"[^0-9]", "", timestamp)[:14]
        # Format as YYYYMMDD_HHMMSS
        if len(run_id) >= 14:
            run_id = run_id[:8] + "_" + run_id[8:14]

        # Check if already exists
        existing = db.execute(
            "SELECT 1 FROM runs WHERE timestamp = ? AND mode = ? AND model = ?",
            (timestamp, mode, model)
        ).fetchone()
        if existing:
            continue

        db.execute("""
            INSERT INTO runs (
                timestamp, run_id, mode, model,
                turns_used,
                input_tokens, cached_tokens, uncached_tokens,
                output_tokens, total_tokens, cache_hit_pct,
                uncached_cost, cached_cost, output_cost,
                real_cost, nocache_cost, savings, savings_pct,
                real_cost_inr, duration_sec, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            timestamp, run_id, mode, model,
            row[col("Turns Used")],
            row[col("Input Tokens")],
            row[col("Cached Tokens")],
            row[col("Uncached Tokens")],
            row[col("Output Tokens")],
            row[col("Total Tokens")],
            row[col("Cache Hit %")],
            row[col("Uncached Cost ($)")],
            row[col("Cached Cost ($)")],
            row[col("Output Cost ($)")],
            row[col("Real Cost ($)")],
            row[col("No-Cache Cost ($)")],
            row[col("Savings ($)")],
            row[col("Savings %")],
            row[col("Real Cost (₹)")],
            row[col("Duration (sec)")],
            row[col("Status")],
            row[col("Notes")],
        ))
        inserted += 1

    wb.close()
    db.commit()
    return inserted


# ── 2. Enrich runs with data from output .txt files ──────────────────────────

def enrich_runs_from_txt(db: sqlite3.Connection) -> int:
    """Read output_*.txt headers to fill in url, app_name, version, turns_max,
    rolling_summary — fields that Excel doesn't have.

    Returns count of rows updated.
    """
    updated = 0

    for run_dir in RUN_DIRS:
        if not run_dir.exists():
            continue
        for txt_file in run_dir.glob("output_*.txt"):
            # Extract run_id from filename: output_v2_poc_20260311_164518.txt
            m = re.search(r"(\d{8}_\d{6})", txt_file.name)
            if not m:
                continue
            file_run_id = m.group(1)

            # Parse the header block (first ~12 lines before the === separator)
            header = _parse_txt_header(txt_file)
            if not header:
                continue

            # Determine version
            version = "v2" if "v2" in txt_file.name or "v2" in header.get("Version", "") else "v1"

            # Parse "19 / 120" -> turns_used=19, turns_max=120
            turns_str = header.get("Turns", "")
            turns_max = None
            if "/" in turns_str:
                parts = turns_str.split("/")
                try:
                    turns_max = int(parts[1].strip())
                except (ValueError, IndexError):
                    pass

            # Find matching run in DB (by run_id pattern or close timestamp)
            db.execute("""
                UPDATE runs SET
                    url = COALESCE(url, ?),
                    app_name = COALESCE(app_name, ?),
                    version = COALESCE(version, ?),
                    turns_max = COALESCE(turns_max, ?),
                    rolling_summary = COALESCE(rolling_summary, ?)
                WHERE run_id LIKE ?
            """, (
                header.get("URL"),
                header.get("App"),
                version,
                turns_max,
                header.get("Rolling Summary"),
                f"%{file_run_id}%",
            ))
            if db.total_changes > updated:
                updated = db.total_changes

    db.commit()
    return updated


def _parse_txt_header(txt_file: Path) -> dict:
    """Extract key: value pairs from the header block of an output .txt file."""
    header = {}
    try:
        text = txt_file.read_text(errors="replace")
    except OSError:
        return header

    for line in text.splitlines()[:15]:
        if "=" * 10 in line:
            break
        if ":" in line:
            key, _, val = line.partition(":")
            header[key.strip()] = val.strip()
    return header


# ── 3. Sync turns + actions from turns_*.json ────────────────────────────────

def sync_turns(db: sqlite3.Connection) -> tuple[int, int]:
    """Read all turns_*.json files and insert into turns + actions tables.

    Returns (turns_inserted, actions_inserted).
    """
    turns_inserted = 0
    actions_inserted = 0

    for run_dir in RUN_DIRS:
        if not run_dir.exists():
            continue
        for json_file in run_dir.glob("turns_*.json"):
            # Extract run_id from filename
            m = re.search(r"(\d{8}_\d{6})", json_file.name)
            if not m:
                continue
            file_run_id = m.group(1)

            # Find the matching run_id in the DB
            row = db.execute(
                "SELECT run_id FROM runs WHERE run_id LIKE ?",
                (f"%{file_run_id}%",)
            ).fetchone()
            if not row:
                # No matching run in DB — skip (Excel might not have this run)
                continue
            run_id = row[0]

            # Check if turns already synced for this run
            existing = db.execute(
                "SELECT 1 FROM turns WHERE run_id = ? LIMIT 1", (run_id,)
            ).fetchone()
            if existing:
                continue

            # Load and parse JSON
            try:
                data = json.loads(json_file.read_text())
            except (json.JSONDecodeError, OSError):
                continue

            turns_list = data.get("turns", [])
            for turn in turns_list:
                turn_num = turn.get("turn", 0)

                db.execute("""
                    INSERT INTO turns (
                        run_id, turn_number,
                        input_tokens, cached_tokens, uncached_tokens,
                        output_tokens, new_tokens, cache_hit_ratio,
                        thought
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    run_id, turn_num,
                    turn.get("input_tokens"),
                    turn.get("cached_tokens"),
                    turn.get("uncached_tokens"),
                    turn.get("output_tokens"),
                    turn.get("new_tokens"),
                    turn.get("cache_hit_ratio"),
                    turn.get("thought"),
                ))
                turns_inserted += 1

                # Insert actions for this turn
                for action in turn.get("actions", []):
                    db.execute("""
                        INSERT INTO actions (run_id, turn_number, tool, target)
                        VALUES (?, ?, ?, ?)
                    """, (
                        run_id, turn_num,
                        action.get("tool", ""),
                        action.get("target", ""),
                    ))
                    actions_inserted += 1

    db.commit()
    return turns_inserted, actions_inserted


# ── 4. Sync fields + failures from StateTracker run_*.json ───────────────────

def sync_state_tracker(db: sqlite3.Connection) -> tuple[int, int]:
    """Read state/runs/run_*.json files and insert fields + failures.

    Returns (fields_inserted, failures_inserted).
    """
    fields_inserted = 0
    failures_inserted = 0

    if not STATE_DIR.exists():
        return 0, 0

    for json_file in STATE_DIR.glob("run_*.json"):
        try:
            data = json.loads(json_file.read_text())
        except (json.JSONDecodeError, OSError):
            continue

        run_id = data.get("run_id", "")
        if not run_id:
            continue

        # Skip if this run_id doesn't exist in the runs table (FK would fail)
        has_run = db.execute(
            "SELECT 1 FROM runs WHERE run_id = ? OR run_id LIKE ?",
            (run_id, f"%{run_id}%")
        ).fetchone()
        if not has_run:
            continue

        # Check if fields already synced for this run
        existing = db.execute(
            "SELECT 1 FROM fields WHERE run_id = ? LIMIT 1", (run_id,)
        ).fetchone()
        if existing:
            continue

        # Sections → fields
        for section in data.get("sections", []):
            section_name = section.get("name", "")
            for field in section.get("fields", []):
                db.execute("""
                    INSERT INTO fields (
                        run_id, section_name, field_name, uid,
                        action, xpath, status, value, retries, error
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    run_id, section_name,
                    field.get("name", ""),
                    field.get("uid", ""),
                    field.get("action", ""),
                    field.get("xpath", ""),
                    field.get("status", ""),
                    field.get("value"),
                    field.get("retries", 0),
                    field.get("error"),
                ))
                fields_inserted += 1

        # Failures
        for failure in data.get("failures", []):
            db.execute("""
                INSERT INTO failures (
                    run_id, section_name, field_name, action,
                    error, diagnosis, evidence, failure_timestamp
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                run_id,
                failure.get("section", ""),
                failure.get("field", ""),
                failure.get("action", ""),
                failure.get("error", ""),
                failure.get("diagnosis", ""),
                json.dumps(failure.get("evidence", {})),
                failure.get("timestamp", ""),
            ))
            failures_inserted += 1

    db.commit()
    return fields_inserted, failures_inserted


# ── 5. Show stats ────────────────────────────────────────────────────────────

def show_stats(db: sqlite3.Connection) -> None:
    """Print what's currently in the database."""
    print("\n  DATABASE CONTENTS")
    print(f"  {'='*40}")

    tables = ["runs", "turns", "actions", "fields", "failures", "test_cases"]
    for table in tables:
        count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table:<15} {count:>6} rows")

    print(f"\n  {'─'*40}")

    # Run summary
    row = db.execute("""
        SELECT
            COUNT(*) as runs,
            SUM(real_cost) as total_cost,
            SUM(total_tokens) as total_tokens,
            AVG(cache_hit_pct) as avg_cache,
            MIN(timestamp) as first_run,
            MAX(timestamp) as last_run
        FROM runs
    """).fetchone()

    if row and row[0] > 0:
        print(f"  Total runs:    {row[0]}")
        print(f"  Total cost:    ${row[1]:.4f}" if row[1] else "  Total cost:    $0")
        print(f"  Total tokens:  {row[2]:,.0f}" if row[2] else "  Total tokens:  0")
        print(f"  Avg cache hit: {row[3]:.1f}%" if row[3] else "  Avg cache hit: N/A")
        print(f"  First run:     {row[4]}")
        print(f"  Last run:      {row[5]}")

    # Tool usage breakdown
    tools = db.execute("""
        SELECT tool, COUNT(*) as uses
        FROM actions
        GROUP BY tool
        ORDER BY uses DESC
        LIMIT 10
    """).fetchall()
    if tools:
        print(f"\n  TOP TOOLS")
        print(f"  {'─'*40}")
        for tool, count in tools:
            print(f"  {tool:<25} {count:>6} calls")

    # Cost by mode
    modes = db.execute("""
        SELECT mode, COUNT(*) as runs, AVG(real_cost) as avg_cost, SUM(real_cost) as total
        FROM runs GROUP BY mode ORDER BY total DESC
    """).fetchall()
    if modes:
        print(f"\n  COST BY MODE")
        print(f"  {'─'*40}")
        print(f"  {'Mode':<15} {'Runs':>5} {'Avg Cost':>10} {'Total':>10}")
        for mode, runs, avg, total in modes:
            avg_str = f"${avg:.4f}" if avg else "$0"
            total_str = f"${total:.4f}" if total else "$0"
            print(f"  {mode:<15} {runs:>5} {avg_str:>10} {total_str:>10}")

    print(f"\n  {'='*40}")
    print(f"  DB file: {DB_PATH}")
    print(f"  DB size: {DB_PATH.stat().st_size / 1024:.1f} KB")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"\n  Syncing to {DB_PATH}")
    print(f"  {'='*40}")

    db = get_db()

    # Step 1: Runs from Excel
    n = sync_runs_from_excel(db)
    print(f"  [runs]       {n} new rows from Excel")

    # Step 2: Enrich with txt headers
    n = enrich_runs_from_txt(db)
    print(f"  [runs]       enriched from .txt headers")

    # Step 3: Turns + actions from JSON
    t, a = sync_turns(db)
    print(f"  [turns]      {t} new rows")
    print(f"  [actions]    {a} new rows")

    # Step 4: Fields + failures from StateTracker
    f, fl = sync_state_tracker(db)
    print(f"  [fields]     {f} new rows")
    print(f"  [failures]   {fl} new rows")

    print(f"\n  Sync complete.")

    # Show stats if requested or always on first sync
    if "--stats" in sys.argv or n > 0 or t > 0:
        show_stats(db)

    db.close()


if __name__ == "__main__":
    main()
